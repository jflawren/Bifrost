# UI (polys/mindep only)
# Author :      Nathan Krueger
# Created       5:00 PM 7/16/15
# Last Updated  3:30 PM 6/4/16
# Version       2.61

import excel
from openpyxl import *
from os.path import exists

output = 0

menu = """
Enter a command:
output   change type of output between shell text and external documentation
polys    polysemy rating for a selection of words
mindep   mindepth of word/s (first result only, see dtree for alternate synsets)
pol_min  runs both polysemy and mindepth analysis of a selection of words
dtree    depth tree of a given word (working on making it neater)

q        quit

command: """

def interface()->(str, int):
    """asks the user what to do and then asks the controller to do it"""
    global output
    while True:
        cmd = input(menu).strip().lower()
        if cmd == 'output':
            output = change_output()
            return (None,None)
        elif cmd == 'polys':
            return ('polys', mwa())
        elif cmd == 'mindep':
            return ('mindep', mwa())
        elif cmd == 'pol_min':
            return ('pol_min', mwa())
        elif cmd == 'dtree':
            return ('dtree', swa())
        elif cmd == 'q':
            print("Goodbye")
            return ('quit', None)
        else:
            print("Invalid command, please try again")
    return

def change_output()->int:
    '''allows user to change the programs output to text and/or excel file'''
    while True:
        output_type = input("Change output to (excel, default/text/shell text, both): ").strip().lower()
        if output_type in ['text', 'shell text', 'default']:
            return 0
        elif output_type == 'excel':
            return 1
        elif output_type == 'both':
            return 2
        else:
            print("Invalid output type, please select from: excel, text, or both")

def swa()->str:
    """run an analysis on a signle word"""
    return input("Please enter a word to analyze: ").strip().lower()

def mwa()->[str]:
    """run an analysis on several words"""
    while True:
        print("Word sources: text, excel, manual (.txt, .xls/.xlsx, type into console)")
        word_source = input("Please enter a source for the words to analyze: ").strip().lower()
        if word_source not in ['text', 'excel', 'manual']:
            print("This is not a valid word source")
            continue
        break
    if word_source == 'manual':
        words = input("Please enter a series of words to analyze seperated only by spaces: ").strip().lower().split()
        return words
    if word_source == 'excel':
        return excel_words()
    if word_source == 'text':
        return text_words()

def text_words()->[str]:
    '''takes a text file and returns a list of words'''
    print("""Please enter the fileid of the text file exactly as it is
(including the extension), please ensure that the file is present in
the current folder and that all words seperated by only whitespace
""")
    while True:
        file_name = input("fileid of text file: ").strip()
        try:
            file = open(file_name).read().splitlines()
            break
        except:
            print("This text file is not available, please make sure that you typed it correctly")
            continue
    result = []
    for line in file:
        line = line.strip().lower()
    if len(file[0]) == 1:
        result = file
        return result
    for line in file:
        for word in line.split():
            result.append(word)
    return result

def excel_words()->[str]:
    """run an analysis on many words from an excel file"""
    print("""Please enter the fileid of the excel file exactly as it is
(including the extension), please ensure that the file is present in
the current folder and that all words are listed in the first column
""")
    while True:
        file = input("fileid of excel document: ").strip()
        sheet = input("Please enter the sheet name to check (by default, Excel uses Sheet1): ").strip()
        try:
            file = load_workbook(file)
            sheet = file.get_sheet_by_name(sheet)
            break
        except:
            print("This excel document or sheet is not available, please make sure that you \ntyped it correctly")
            continue
    result = []
    for pos in range(len(sheet.rows)):
        try:
            result.append(str(sheet.cell(row = pos + 1, column = 1).value))
            if dual:
                result.append(str(sheet.cell(row = pos + 1, column = 2).value))
        except:
            pass
    return result

def output_data(value)->None:
    '''selects between output styles and call the correct function/s'''
    if output != 1:
        print_data(value)
    if output != 0:
        data_to_file(value)

def data_to_file(value)->None:
    '''outputs the data as a file of specified name and type'''
    print("""
Warning: for the moment this program cannot append to files,
only create and overwrite them, please be careful about using existing files
You can specify a different sheet name to use if you wish to add to a file

Also, please choose a file type of .xlsx (.xls is not currently supported)
and be sure that the file is not open in another window.
          """)
    while True:
        file_name = input("Please enter the name of the file you would to create, including the extension: ")
        print("\n{}\n".format(file_name))
        sure = input("Are you sure this is the file name you wish to use (y/n)").strip().lower()
        if sure not in ['y', 'yes']:
            continue
        if file_name.split('.')[1] != 'xlsx':
            print("This file type is not currently supported")
            continue
        sheet_name = input("Please enter a sheet name: ").strip()
        break
    file_from_data(value, file_name, sheet_name)
    return

def file_from_data(value, file_name: str, sheet_index: int)->None:
    '''creates an output file using the given name'''
    data, function = value
    if function in ['newc','dtree']:
        return
    excel.file_setup(data, file_name, sheet_index)
    return

def print_data(value)->None:
    """prints the data"""
    #can probably shorten this if I try
    data, function = value
    if function == 'polys' or function == 'pol_min':
        print("\nNumber of defintions per part of speech for each word:")
        print("\nWord                Noun  Adj  SatAdj  Adv  Verb")
        for word in data:
            print("{:18}{:6}{:5}{:8}{:5}{:6}".format(word[0],word[1],word[2],word[3],word[4],word[5]))
    if function == 'pol_min':
        print("""\nMin depth of the first, or most common definition for each part of
speech of each word; -1 signifies that no defintion of that type was found:""")
        print("\nWord  Min depth as: Noun  Adj  SatAdj  Adv  Verb")
        for word in data:
            print("{:18}{:6}{:5}{:8}{:5}{:6}".format(word[0],word[6],word[7],word[8],word[9],word[10]))
    if function == 'mindep':
        print("""\nMin depth of the first, or most common definition for each part of
speech of each word; -1 signifies that no defintion of that type was found:""")
        print("\nWord  Min depth as: Noun  Adj  SatAdj  Adv  Verb")
        for word in data:
            print("{:18}{:6}{:5}{:8}{:5}{:6}".format(word[0],word[6],word[7],word[8],word[9],word[10]))
    if function == 'dtree':
        print_dtree(data)
    return

def print_dtree(data):
    '''prints dtree data'''
    from pprint import pprint
    depth = lambda L: isinstance(L, list) and max(map(depth, L))+1
    d_count = 0
    if (len(data[1]) > 1):
        print("Multiple definitions for the word {} detected: \n".format(data[0]))
        print("Defintions:")
        for def_index in range(len(data[1])):
            if data[2][def_index] == 'n':
                print("{}: noun: {}".format(def_index, data[1][def_index]))
                d_count = d_count + 1
        while True:
            action = input('\nWhich dtree would you like (choose an index or say "all"): ')
            if action == 'all':
                print("\nNote: multiple entires on the same line are equivalent")
                for def_index in range(d_count):
                    print("\nDtree of '{}' with depth {} as defined as: {}".format(data[0],depth(data[3][def_index]) - 1,data[1][def_index]))
                    pprint(data[3][def_index])
                return
            else:
                try:
                    action = int(action)
                    if (action >= d_count):
                        raise fail
                    dtree = data[3][int(action)]
                    #in this order to make the note only appear if it works
                    print("\nNote: multiple entires on the same line are equivalent")
                    print("\nDtree of '{}' with depth {} as defined as: {}".format(data[0],depth(dtree) - 1,data[1][action]))
                    pprint(dtree)
                    return
                except:
                    print('This index is either invalid or not "all"')
    elif (len(data[1]) == 1):
        print('Only 1 defintion of the word "{}" wss found: '.format(data[0]))
        print("\nNote: multiple entires on the same line are equivalent")
        print("\nDtree of '{}' as defined as: {}".format(data[0],data[1][0]))
        pprint(data[3][0])
    else:
        print('No defintions for the word "{}" were found...'.format(data[0]))
    return
