# Excel (polys/mindep only)
# Author :      Nathan Krueger
# Created       11:45 AM 8/9/15
# Last Updated  3:30 PM 6/4/16
# Version       2.8

from openpyxl import *

file = None
sheet = None

def file_setup(data, file_name: str, sheet_name = '', sheet_index = 0)->None:
    '''sets up the files header row'''
    if type(data) != list or type(data[0]) != list or len(data[0]) == 0:
        print("\n\n\nSomething happened to the data... Either you screwed up or I did. Double check and if you think I did then this is a bug, write down what you did and send it to me.\n\n\n")
        return
    global file, sheet
    try: #file exists, add a new sheet
        file = load_workbook(file_name)
        sheet = file.create_sheet(title = sheet_name)
    except: #file does not exist, rename first sheet
        try:
            file = Workbook()
            sheet = file.active
            sheet.title = sheet_name
        except:
            print("Err: make sure that the file is closed if it exists!")
            file_setup(data, file_name, sheet_name)
    polys_mindep_setup(data)
    file.save(file_name)
    return

def polys_mindep_setup(data)->None:
    '''sets up a polys data spreadsheet'''
    sheet.cell(row = 1, column = 1).value = 'Word'
    sheet.cell(row = 1, column = 2).value = 'polys.noun'
    sheet.cell(row = 1, column = 3).value = 'polys.adj'
    sheet.cell(row = 1, column = 4).value = 'polys.sat_adj'
    sheet.cell(row = 1, column = 5).value = 'polys.adv'
    sheet.cell(row = 1, column = 6).value = 'polys.verb'
    sheet.cell(row = 1, column = 7).value = 'mindep.noun'
    sheet.cell(row = 1, column = 8).value = 'mindep.adj'
    sheet.cell(row = 1, column = 9).value = 'mindep.sat_adj'
    sheet.cell(row = 1, column = 10).value = 'mindep.adv'
    sheet.cell(row = 1, column = 11).value = 'mindep.verb'
    write_polys_mindep(data)
    return

def write_polys_mindep(data)->None:
    '''writes polysemy data to the file'''
    index = 2
    for word in data:
        sheet.cell(row = index, column = 1).value = word[0]
        sheet.cell(row = index, column = 2).value = word[1]
        sheet.cell(row = index, column = 3).value = word[2]
        sheet.cell(row = index, column = 4).value = word[3]
        sheet.cell(row = index, column = 5).value = word[4]
        sheet.cell(row = index, column = 6).value = word[5]
        sheet.cell(row = index, column = 7).value = word[6]
        sheet.cell(row = index, column = 8).value = word[7]
        sheet.cell(row = index, column = 9).value = word[8]
        sheet.cell(row = index, column = 10).value = word[9]
        sheet.cell(row = index, column = 11).value = word[10]
        index += 1
    return
